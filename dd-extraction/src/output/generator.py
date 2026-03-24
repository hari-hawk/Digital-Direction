"""
Excel output generator.
Produces Dynamics-ready inventory file with 3-tier headers, dropdowns, and formatting.
Updated for 54-column template format.
"""
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.mapping.schema import (
    INVENTORY_SCHEMA, SECTION_AREAS, COLUMN_NAMES,
    SERVICE_TYPES, CHARGE_TYPES, SCU_CODES, STATUS_VALUES,
    MONTH_TO_MONTH_VALUES, InventoryRow,
)

# Styling constants
HEADER_FONT = Font(bold=True, size=11)
SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
TIER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Checklist items from template (30 items)
CHECKLIST_ITEMS = [
    "Check whether S record missing  and only C record is entered for a service",
    "Loc Z addresses",
    "Sub total mismatch - S record total to match with the C records",
    "Service Address Not Available, check with source file once",
    "Reconcile the  total with Monthly report / invoice",
    "Service address normalization along with zip to be done",
    "One TN with 2 service types to be  checked through pivot",
    "1SXWE Usoc  service type to be centrex (applicable for Verizon vendor)",
    "DID in Service type 2 & Primary service  in service type1",
    "Bandwidth should be captured on both S & C record .  The Access and Upload speed would be Symmetrical in case of Ethernet , MPLS, DIA and Optical Circuit, DS1, DS3",
    "Account level charges to be captured in inventory",
    "Account/ Sub Acct/ BTN/ Carrier acts to be validated with source file",
    "Billing Names should not be blank",
    "Zero line POTS charges.",
    "Additional items details to be updated from CSR's",
    "Formatting of $ in Monthly cost & Cost each",
    "Formatting of Excel file to be uniform across",
    "Phone Number Format should be special",
    "Inventory notes to be validated  for anything missing/TBD",
    "If TBD, add details in service type 2 for further research for DD1",
    "3 Addresses - 3rd address to be mentioned in service type 2 (incase found any can be  flagged)",
    'Fill all blanks with "No Fill" (Applicable for NYC for now, incase of Vendor-Verizon)',
    "Ethernet service should have Access & Upload speed",
    "Dark Fiber can be completed without Bandwidth",
    "Recheck for blanks. If anything missing it cannot be tagged as inventory complete",
    "Check for one circuit billing in 2 accounts and update service type 2",
    'For RJ21X description we can use Inside Wiring for Service Type (If Applicable)',
    'For Entrance Bridge, Additional item use " Conferencing" as service type',
    'For all Managed security and Managed services use " Managed Services" as the service type',
    "The AT&T HSI services are always Month to Month",
]

# Column explanations from template
COLUMN_EXPLANATIONS = {
    "Status": "Complete it all information is obtained the record is complete.\nPending if all information is not obtained and\\or building the inventory is ongoing.",
    "Inventory creation questions, concerns, or notes.": "Used to show what the inventory is still missing or other helpful information.",
    "*Contract Info received": "Contract information and the actual documents is hard to obtain. This column refers to the current status of the work to obtain this information.",
    "Invoice File Name": "The exact name of the invoice file the inventory is based on.",
    "Files Used For Inventory": "The exact name of any other files used to build the inventory besides the invoice.",
    "Service Type": "See Dropdown tab for the Service Types we use.",
    "Service Type 2": "If we feel additional Service Type information is helpful, we show here.",
}

# Dropdown explanation pairs from template
STATUS_EXPLANATIONS = [
    "Used when all possible information is obtained",
    "Used when the invoice provided is not applicable to our needs",
    "Used when we are working to obtain carrier information or finalize",
    "Used when DD decides we will no longer pursue the missing information",
]

CHARGE_TYPE_EXPLANATIONS = [
    "Monthly reoccuring charges",
    "Non reoccuring charges",
    "Other credit and charges",
    "Self explanatory",
    "Self explanatory",
    "Self explanatory",
    "Self explanatory",
]

SCU_EXPLANATIONS = [
    "Used to show the cost of the service in totality without taxes, sucharges, other credits\\charges.",
    "Used to show the components that make up the S record.",
    "Usage fees.",
    "Taxes, Surcharges, Other Credit and Charges",
]

# Notes column paired with Service Types
SERVICE_TYPE_NOTES = [
    "Keep this for unknowns/Unable to confirm",
    "To be used if no designation of MPLS or Internet can be discerned and Ethernet verbiage is used by the carrier.",
    "Hosted voip is only Cloud IP-PBX.",
    "Dynamic IP",
    "Keep MPLS, use MPLS for VPLS, IPVPN & MPLS service types",
    "Concurrent Call Paths w/ any number of DIDs. Used with a phone system like a Cisco Call Manager. IP Flex is another example",
    "Ideally we can inventory useage as part of a particular service. In some cases the carrier information only indicates it is usage without enough information to correlate it to a specific number. In other cases it might not be a good use of time to correlate every penny of usage to a specific number. Type of useage can be defined in the Component or Feature Name column.",
    "N/A for 2024",
    "N/A for 2024",
    "Duplicate of VTN",
    "1::1 phone num/ CCP; no phone system required",
    "Bought in builk to run over MPLS, Ethernet, IP VPN",
    "Cradlepoint",
    "Use for Fixed wireless, not Cradle points",
]


def generate_inventory_excel(
    rows: list[InventoryRow],
    output_path: Path,
    carrier_name: str = "Charter Communications",
) -> Path:
    """
    Generate the Dynamics-ready Excel inventory file (54-column format).

    Creates sheets matching the template:
    - Baseline: 3-row header + data rows (54 columns)
    - Columns Explained if Needed
    - Dropdowns: 9 dropdown lists with explanations
    - Empty Template: headers only, no data
    - Removed Options
    - Instructions
    - Account Tracking
    - hiddenSheet
    - Checklist: 30 QA items
    - TF, DID
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- 1. Baseline Sheet ---
    ws = wb.active
    ws.title = "Baseline"
    _write_baseline_sheet(ws, rows)

    # --- 2. Columns Explained if Needed ---
    ws_expl = wb.create_sheet("Columns Explained if Needed")
    _write_explanations_sheet(ws_expl)

    # --- 3. Dropdowns Sheet ---
    ws_dd = wb.create_sheet("Dropdowns")
    _write_dropdowns_sheet(ws_dd)

    # --- 4. Empty Template (headers only, no data) ---
    ws_et = wb.create_sheet("Empty Template")
    _write_baseline_sheet(ws_et, [])  # same headers, zero data rows

    # --- 5. Removed Options ---
    ws_ro = wb.create_sheet("Removed Options")
    ws_ro.cell(row=1, column=1, value="Removed Options").font = HEADER_FONT

    # --- 6. Instructions ---
    ws_instr = wb.create_sheet("Instructions ")
    ws_instr.cell(row=1, column=1, value="Instructions").font = HEADER_FONT

    # --- 7. Account Tracking ---
    _write_template_sheet(wb, "Account Tracking",
        ["Carrier Account Number", "Carrier", "Current Monthly Charges",
         "Services", "Bill Date", "Contract Begin", "Contract End", "Notes"])

    # --- 8. hiddenSheet ---
    ws_hidden = wb.create_sheet("hiddenSheet")
    ws_hidden.sheet_state = "hidden"

    # --- 9. Checklist ---
    ws_cl = wb.create_sheet(" Checklist")
    _write_checklist_sheet(ws_cl)

    # --- 10. TF, DID ---
    ws_tf = wb.create_sheet("TF, DID")
    _write_tf_did_sheet(ws_tf)

    # --- Apply data validation to Baseline ---
    _apply_data_validation(wb["Baseline"], len(rows))

    wb.save(str(output_path))
    return output_path


def _write_baseline_sheet(ws, rows: list[InventoryRow]):
    """Write the Baseline sheet with 3-tier header structure (54 columns)."""
    num_cols = len(INVENTORY_SCHEMA)

    # --- Row 1: Section Area headers (merged cells) ---
    for section_name, (start_letter, end_letter) in SECTION_AREAS.items():
        start_idx = _letter_to_col_idx(start_letter)
        end_idx = _letter_to_col_idx(end_letter)

        ws.merge_cells(
            start_row=1, start_column=start_idx,
            end_row=1, end_column=end_idx,
        )
        cell = ws.cell(row=1, column=start_idx, value=section_name)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(horizontal="center")

    # --- Row 2: Requirement Tier flags ---
    for i, col_def in enumerate(INVENTORY_SCHEMA):
        cell = ws.cell(row=2, column=i + 1, value=col_def.requirement_tier)
        cell.font = Font(italic=True, size=9)
        cell.fill = TIER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # --- Row 3: Column Names ---
    for i, col_def in enumerate(INVENTORY_SCHEMA):
        cell = ws.cell(row=3, column=i + 1, value=col_def.name)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True)

    # --- Row 4+: Data ---
    for row_idx, inv_row in enumerate(rows, start=4):
        row_dict = inv_row.to_row_dict()
        for col_idx, col_def in enumerate(INVENTORY_SCHEMA):
            value = row_dict.get(col_def.name)
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
            cell.border = THIN_BORDER

    # --- Formatting ---
    ws.freeze_panes = "A4"

    for i, col_def in enumerate(INVENTORY_SCHEMA):
        col_letter = get_column_letter(i + 1)
        max_width = max(len(col_def.name), 12)
        ws.column_dimensions[col_letter].width = min(max_width + 2, 30)


def _write_explanations_sheet(ws):
    """Write the Columns Explained if Needed sheet."""
    ws.cell(row=1, column=1, value="Column Name").font = HEADER_FONT
    ws.cell(row=1, column=2, value="Explanation").font = HEADER_FONT
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 80

    row = 2
    for col_name, explanation in COLUMN_EXPLANATIONS.items():
        ws.cell(row=row, column=1, value=col_name)
        ws.cell(row=row, column=2, value=explanation)
        row += 1


def _write_dropdowns_sheet(ws):
    """Write all 9 dropdown validation lists matching the template."""
    col = 1

    # --- Column A: New Motor Service Type (86 values) ---
    ws.cell(row=1, column=col, value="New Motor Service Type").font = HEADER_FONT
    for i, val in enumerate(SERVICE_TYPES, start=2):
        ws.cell(row=i, column=col, value=val)
    col += 1

    # --- Column B: Notes (service type notes) ---
    ws.cell(row=1, column=col, value="Notes").font = HEADER_FONT
    for i, val in enumerate(SERVICE_TYPE_NOTES, start=2):
        ws.cell(row=i, column=col, value=val)
    col += 1

    # --- Column C: Status ---
    ws.cell(row=1, column=col, value="Status").font = HEADER_FONT
    for i, val in enumerate(STATUS_VALUES, start=2):
        ws.cell(row=i, column=col, value=val)
    col += 1

    # --- Column D: Status Explanation ---
    ws.cell(row=1, column=col, value="Status Explanation").font = HEADER_FONT
    for i, val in enumerate(STATUS_EXPLANATIONS, start=2):
        ws.cell(row=i, column=col, value=val)
    col += 1

    # --- Column E: Charge Type ---
    ws.cell(row=1, column=col, value="Charge Type").font = HEADER_FONT
    for i, val in enumerate(CHARGE_TYPES, start=2):
        ws.cell(row=i, column=col, value=val)
    col += 1

    # --- Column F: Charge Type Explanation ---
    ws.cell(row=1, column=col, value="Charge Type Explanation").font = HEADER_FONT
    for i, val in enumerate(CHARGE_TYPE_EXPLANATIONS, start=2):
        ws.cell(row=i, column=col, value=val)
    col += 1

    # --- Column G: Service or Component Options ---
    ws.cell(row=1, column=col, value="Service or Component Options").font = HEADER_FONT
    for i, val in enumerate(SCU_CODES, start=2):
        ws.cell(row=i, column=col, value=val)
    col += 1

    # --- Column H: Service or Component Options Explained ---
    ws.cell(row=1, column=col, value="Service or Component Options Explained").font = HEADER_FONT
    for i, val in enumerate(SCU_EXPLANATIONS, start=2):
        ws.cell(row=i, column=col, value=val)
    col += 1

    # --- Column I: Currently Month-to-Month ---
    ws.cell(row=1, column=col, value="Currently Month-to-Month").font = HEADER_FONT
    for i, val in enumerate(MONTH_TO_MONTH_VALUES, start=2):
        ws.cell(row=i, column=col, value=val)

    # Adjust widths for all 9 columns
    for c in range(1, 10):
        ws.column_dimensions[get_column_letter(c)].width = 30


def _write_checklist_sheet(ws):
    """Write the Checklist sheet with 30 QA items."""
    headers = ["Checklist Item", "Agent - Yes/No", "QA - Yes/No"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.border = THIN_BORDER

    for row_idx, item in enumerate(CHECKLIST_ITEMS, start=2):
        ws.cell(row=row_idx, column=1, value=item).border = THIN_BORDER
        ws.cell(row=row_idx, column=2).border = THIN_BORDER
        ws.cell(row=row_idx, column=3).border = THIN_BORDER

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18


def _write_template_sheet(wb, sheet_name: str, headers: list):
    """Create a template sheet with headers only."""
    ws = wb.create_sheet(sheet_name)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="27272A")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="left")
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


def _write_tf_did_sheet(ws):
    """Write TF, DID sheet."""
    headers = ["Vendor", "Account", "Subaccount", "DID", "TN"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER


def _apply_data_validation(ws, num_rows: int):
    """Apply dropdown data validation to the Baseline sheet (54-column layout)."""
    if num_rows == 0:
        return

    last_row = num_rows + 3  # 3 header rows + data

    # Status (column A = 1)
    dv_status = DataValidation(
        type="list",
        formula1=f"Dropdowns!$C$2:$C${len(STATUS_VALUES) + 1}",
        allow_blank=True,
    )
    dv_status.error = "Invalid Status"
    dv_status.errorTitle = "Status"
    ws.add_data_validation(dv_status)
    dv_status.add(f"A4:A{last_row}")

    # Service Type (column V = 22)
    dv_service = DataValidation(
        type="list",
        formula1=f"Dropdowns!$A$2:$A${len(SERVICE_TYPES) + 1}",
        allow_blank=True,
    )
    dv_service.error = "Invalid Service Type"
    dv_service.errorTitle = "Service Type"
    ws.add_data_validation(dv_service)
    dv_service.add(f"V4:V{last_row}")

    # Service or Component (column Y = 25)
    dv_scu = DataValidation(
        type="list",
        formula1=f"Dropdowns!$G$2:$G${len(SCU_CODES) + 1}",
        allow_blank=True,
    )
    dv_scu.error = "Invalid Service or Component code"
    ws.add_data_validation(dv_scu)
    dv_scu.add(f"Y4:Y{last_row}")

    # Charge Type (column AD = 30)
    dv_charge = DataValidation(
        type="list",
        formula1=f"Dropdowns!$E$2:$E${len(CHARGE_TYPES) + 1}",
        allow_blank=True,
    )
    dv_charge.error = "Invalid Charge Type"
    ws.add_data_validation(dv_charge)
    dv_charge.add(f"AD4:AD{last_row}")

    # Currently Month-to-Month (column AV = 48)
    dv_mtm = DataValidation(
        type="list",
        formula1=f"Dropdowns!$I$2:$I${len(MONTH_TO_MONTH_VALUES) + 1}",
        allow_blank=True,
    )
    dv_mtm.error = "Invalid value - must be Yes or No"
    ws.add_data_validation(dv_mtm)
    dv_mtm.add(f"AV4:AV{last_row}")


def _letter_to_col_idx(letter: str) -> int:
    """Convert Excel column letter to 1-based index."""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result
