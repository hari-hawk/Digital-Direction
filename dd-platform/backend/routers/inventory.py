from __future__ import annotations
"""Inventory router: read and filter inventory data with multi-sheet support."""
from pathlib import Path

import pandas as pd
import openpyxl
from fastapi import APIRouter, Request
from fastapi.responses import FileResponse

from services.analytics_service import get_inventory_rows, load_inventory

router = APIRouter(tags=["inventory"])

# Cache for sheet data
_sheet_cache = {}  # type: dict

# In-memory store for row status/comments (POC — replace with Neon DB later)
_row_status_store: dict[str, dict[int, dict]] = {}  # project_id -> { row_index -> { status, comment } }

# Required fields used to compute per-row accuracy
_REQUIRED_FIELDS = [
    "Carrier", "Carrier Account Number", "Service Type", "Charge Type",
    "Service or Component", "Billing Name", "Service Address", "City",
    "State", "Zip", "Phone Number", "Carrier Circuit Number",
    "Monthly Recurring Cost", "Component or Feature Name",
]


def _compute_row_accuracy(row_data: dict) -> int:
    """Compute accuracy 0-100 based on how many required fields are populated."""
    populated = 0
    total = 0
    for field in _REQUIRED_FIELDS:
        # Match case-insensitively against the row data keys
        matched_key = next((k for k in row_data if k.strip().lower() == field.lower()), None)
        if matched_key is None:
            # Also try partial match
            matched_key = next((k for k in row_data if field.lower() in k.lower()), None)
        total += 1
        if matched_key:
            val = row_data.get(matched_key)
            if val is not None and str(val).strip() not in ("", "nan", "None"):
                populated += 1
    return round((populated / total) * 100) if total > 0 else 0


def _get_row_status(project_id: str, row_index: int, accuracy: int) -> str:
    """Get status for a row: user-set status takes priority, else auto-derived."""
    store = _row_status_store.get(project_id, {})
    entry = store.get(row_index)
    if entry and entry.get("status"):
        return entry["status"]
    # Auto-derive from accuracy
    if accuracy >= 90:
        return "completed"
    return "need_review"


def _get_source_files(proj: dict, carrier_name: str = "", row_data: dict = None) -> list[dict]:
    """Get source documents specific to this row based on invoice filename, account number, and carrier.

    Matching priority:
    1. Exact invoice filename match (from Invoice File Name column)
    2. Files Used For Inventory column values
    3. Carrier Account Number appears in filename
    4. Fallback: carrier name in folder/filename
    """
    input_dir = Path(proj.get("input_dir", ""))
    if not input_dir.exists():
        return []

    # Extract row identifiers for matching
    row_data = row_data or {}
    invoice_file = str(row_data.get("Invoice File Name", "") or "").strip()
    files_used = str(row_data.get("Files Used For Inventory", "") or "").strip()
    carrier = str(row_data.get("Carrier", carrier_name) or carrier_name or "").strip()
    account_num = str(row_data.get("Carrier Account Number", "") or "").strip()

    # Build search terms from row data
    search_terms = set()
    if invoice_file and invoice_file.lower() not in ("nan", "none", ""):
        clean = invoice_file.replace(".pdf", "").replace(".PDF", "").strip()
        search_terms.add(clean.lower())
    if files_used and files_used.lower() not in ("nan", "none", ""):
        for part in files_used.split(","):
            clean = part.strip().replace(".pdf", "").replace(".PDF", "").strip()
            if clean and clean.lower() not in ("nan", "none"):
                search_terms.add(clean.lower())

    carrier_lower = carrier.lower()
    carrier_words = carrier_lower.split()[:2] if carrier_lower else []
    acct_clean = account_num.lower().replace(" ", "").replace("-", "") if account_num else ""

    # Scan all files in input directory
    scored: list[tuple] = []
    for category in ["Invoices", "Contracts", "Carrier Reports, Portal Data, ETC", "CSRs"]:
        cat_dir = input_dir / category
        if not cat_dir.exists():
            continue
        doc_type = ("Invoice" if "invoice" in category.lower() else
                    "Contract" if "contract" in category.lower() else
                    "Report" if "report" in category.lower() else "CSR")

        for f in cat_dir.rglob("*"):
            if not f.is_file() or f.suffix.lower() not in (".pdf", ".xlsx", ".xls", ".csv", ".msg", ".docx"):
                continue

            score = 0
            fstem = f.stem.lower()
            fname_clean = f.name.lower().replace(" ", "").replace("-", "").replace("_", "")

            # Priority 1: Exact invoice filename match
            for term in search_terms:
                term_clean = term.replace(" ", "").replace("-", "").replace("_", "")
                # Must match at least 80% of the search term for it to be "exact"
                min_match_len = max(20, int(len(term_clean) * 0.8))
                if term_clean in fname_clean or (len(term_clean) > 10 and fname_clean.startswith(term_clean[:min_match_len])):
                    score += 100
                    break

            # Priority 2: Account number in filename
            if acct_clean and len(acct_clean) >= 5 and acct_clean in fname_clean:
                score += 50

            # Priority 3: Carrier name in path
            if carrier_words and any(w in fstem for w in carrier_words):
                score += 5

            if score > 0:
                scored.append((score, {
                    "label": f"{doc_type}: {f.name}",
                    "name": f.name,
                    "path": str(f),
                    "format": f.suffix.lower().lstrip("."),
                    "doc_type": doc_type.lower(),
                }))

    # Sort by score descending, deduplicate
    scored.sort(key=lambda x: x[0], reverse=True)
    seen = set()
    result = []
    for score, info in scored:
        if info["name"] not in seen:
            result.append((score, info))
            seen.add(info["name"])

    # If we have exact matches (score >= 100), ONLY return those + account matches
    exact = [info for score, info in result if score >= 50]
    if exact:
        return exact[:5]

    # Otherwise return carrier-level files
    return [info for _, info in result[:5]]


def _resolve_file(proj: dict, source: str) -> str:
    ref = proj.get("reference_file", "") or ""
    if source == "extracted":
        out_dir = proj.get("output_dir", "")
        if out_dir:
            output_dir = Path(out_dir)
            all_carriers = output_dir / "all_carriers_inventory_output.xlsx"
            if all_carriers.exists():
                return str(all_carriers)
            output_files = sorted(output_dir.glob("*_inventory_output.xlsx"))
            if output_files:
                return str(output_files[-1])
        return ref
    return ref


def _load_all_sheets(file_path: str) -> dict:
    """Load all sheet names and metadata from Excel file."""
    if not file_path or not file_path.strip():
        return {"sheets": []}

    if file_path in _sheet_cache:
        return _sheet_cache[file_path]

    path = Path(file_path)
    if not path.exists() or not path.is_file():
        return {"sheets": []}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({
            "name": name.strip(),
            "rows": ws.max_row or 0,
            "cols": ws.max_column or 0,
        })
    wb.close()

    result = {"sheets": sheets}
    _sheet_cache[file_path] = result
    return result


def _load_sheet_data(file_path: str, sheet_name: str, header_row: int = None) -> pd.DataFrame:
    """Load specific sheet from Excel with correct header detection."""
    if not file_path or not file_path.strip():
        return pd.DataFrame()
    path = Path(file_path)
    if not path.exists() or not path.is_file():
        return pd.DataFrame()

    # Baseline uses header row 2 (0-indexed), others use 0
    if sheet_name.strip() == "Baseline":
        header_row = header_row if header_row is not None else 2
    elif sheet_name.strip() in ("Inactive Services", "Windstream Zero Cost Charge", "Migrated Accounts"):
        # These have a similar 3-row header structure
        header_row = header_row if header_row is not None else 0
    else:
        header_row = header_row if header_row is not None else 0

    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    return df


@router.get("/projects/{project_id}/inventory/sheets")
async def list_sheets(project_id: str, request: Request, source: str = "reference"):
    """List all sheets in the inventory file."""
    proj = request.app.state.projects.get(project_id)
    if not proj:
        return {"error": "Project not found"}

    file_path = _resolve_file(proj, source)
    return _load_all_sheets(file_path)


@router.get("/projects/{project_id}/inventory")
async def list_inventory(
    project_id: str,
    request: Request,
    carrier: str = None,
    service_type: str = None,
    charge_type: str = None,
    scu_code: str = None,
    status: str = None,
    review_status: str = None,
    search: str = None,
    sort_by: str = None,
    sort_dir: str = "asc",
    page: int = 1,
    page_size: int = 50,
    source: str = "reference",
    sheet: str = "Baseline",
):
    proj = request.app.state.projects.get(project_id)
    if not proj:
        return {"error": "Project not found"}

    file_path = _resolve_file(proj, source)

    if sheet.strip() == "Baseline":
        # Load full data with standard filters (no pagination yet)
        full_result = get_inventory_rows(
            file_path=file_path,
            carrier=carrier,
            service_type=service_type,
            charge_type=charge_type,
            scu_code=scu_code,
            status=status,
            search=search,
            sort_by=sort_by,
            sort_dir=sort_dir,
            page=1,
            page_size=999999,  # Get all rows for accuracy computation
        )

        # Compute accuracy and review status for EVERY row
        all_rows = full_result.get("rows", [])
        # Compute accuracy and get row-specific source files
        for row in all_rows:
            acc = _compute_row_accuracy(row.get("data", {}))
            row["accuracy"] = acc
            row_review = _get_row_status(project_id, row["row_index"], acc)
            row["status"] = row_review
            # Get source files specific to THIS row (by invoice filename + account number)
            row["source_files"] = _get_source_files(proj, row_data=row.get("data", {}))

        # Apply review_status filter AFTER accuracy computation
        if review_status:
            review_map = {
                "completed": lambda r: r["accuracy"] >= 90,
                "need_review": lambda r: 70 <= r["accuracy"] < 90,
                "critical": lambda r: r["accuracy"] < 70,
            }
            filter_fn = review_map.get(review_status)
            if filter_fn:
                all_rows = [r for r in all_rows if filter_fn(r)]

        # Apply sorting on accuracy/review if requested
        if sort_by == "accuracy":
            all_rows.sort(key=lambda r: r.get("accuracy", 0), reverse=(sort_dir == "desc"))
        elif sort_by == "review" or sort_by == "status":
            order = {"critical": 0, "need_review": 1, "in_progress": 2, "completed": 3}
            if sort_dir == "desc":
                order = {"completed": 0, "in_progress": 1, "need_review": 2, "critical": 3}
            all_rows.sort(key=lambda r: order.get(r.get("status", ""), 99))

        # Now paginate
        total = len(all_rows)
        start = (page - 1) * page_size
        end = start + page_size
        page_rows = all_rows[start:end]

        return {
            "rows": page_rows,
            "total": total,
            "page": page,
            "page_size": page_size,
            "columns": full_result.get("columns", []),
        }

    # Generic sheet loading for non-Baseline sheets
    df = _load_sheet_data(file_path, sheet)
    if df.empty:
        return {"rows": [], "total": 0, "page": page, "page_size": page_size, "columns": []}

    # Apply search filter
    if search:
        mask = pd.Series([False] * len(df))
        for col in df.columns:
            mask |= df[col].astype(str).str.contains(search, case=False, na=False)
        df = df[mask]

    # Apply sorting
    if sort_by and sort_by in df.columns:
        ascending = sort_dir != "desc"
        df = df.sort_values(sort_by, ascending=ascending, na_position="last")

    total = len(df)
    start = (page - 1) * page_size
    end = start + page_size
    page_df = df.iloc[start:end]

    rows = []
    for idx, row in page_df.iterrows():
        row_dict = {}
        for col in df.columns:
            val = row[col]
            if pd.isna(val):
                row_dict[col] = None
            elif hasattr(val, 'item'):
                row_dict[col] = val.item()
            elif hasattr(val, 'isoformat'):
                row_dict[col] = val.isoformat()
            else:
                row_dict[col] = val
        rows.append({"row_index": int(idx), "data": row_dict})

    return {
        "rows": rows,
        "total": total,
        "page": page,
        "page_size": page_size,
        "columns": list(df.columns),
    }


@router.get("/projects/{project_id}/inventory/columns")
async def inventory_columns(project_id: str, request: Request, sheet: str = "Baseline"):
    proj = request.app.state.projects.get(project_id)
    if not proj:
        return {"error": "Project not found"}

    if sheet.strip() == "Baseline":
        df = load_inventory(proj["reference_file"])
    else:
        df = _load_sheet_data(proj["reference_file"], sheet)

    if df.empty:
        return []
    return [{"name": col, "index": i} for i, col in enumerate(df.columns)]


@router.get("/projects/{project_id}/inventory/confidence-summary")
async def confidence_summary(project_id: str, request: Request, source: str = "reference"):
    """Compute confidence summary across all inventory rows."""
    proj = request.app.state.projects.get(project_id)
    if not proj:
        return {"error": "Project not found"}

    empty_result = {
        "total_rows": 0, "high": 0, "medium": 0, "needs_review": 0,
        "high_pct": 0, "medium_pct": 0, "needs_review_pct": 0,
        "high_mrc": 0, "medium_mrc": 0, "needs_review_mrc": 0,
        "extraction_methods": [],
    }

    try:
        file_path = _resolve_file(proj, source)
        df = load_inventory(file_path)
        if df.empty:
            return empty_result
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"Failed to load inventory for confidence: {e}")
        return empty_result

    total = len(df)
    high = 0
    medium = 0
    needs_review = 0
    high_mrc = 0.0
    medium_mrc = 0.0
    needs_review_mrc = 0.0

    mrc_col = next((c for c in df.columns if "monthly recurring" in c.lower()), None)

    for idx, row in df.iterrows():
        row_data = {}
        for col in df.columns:
            val = row[col]
            row_data[col] = None if pd.isna(val) else val
        acc = _compute_row_accuracy(row_data)
        raw_mrc = pd.to_numeric(row.get(mrc_col, 0), errors="coerce") if mrc_col else 0
        mrc_val = float(raw_mrc) if pd.notna(raw_mrc) and not (isinstance(raw_mrc, float) and (raw_mrc != raw_mrc)) else 0.0

        if acc >= 90:
            high += 1
            high_mrc += mrc_val
        elif acc >= 70:
            medium += 1
            medium_mrc += mrc_val
        else:
            needs_review += 1
            needs_review_mrc += mrc_val

    # Extraction method breakdown
    carrier_col = next((c for c in df.columns if c.lower() == "carrier"), None)
    methods = []
    if carrier_col and carrier_col in df.columns:
        for carrier, group in df.groupby(carrier_col):
            if pd.isna(carrier):
                continue
            carrier_rows = len(group)
            carrier_mrc_raw = pd.to_numeric(group[mrc_col], errors="coerce").fillna(0).sum() if mrc_col else 0
            carrier_mrc = float(carrier_mrc_raw) if pd.notna(carrier_mrc_raw) else 0.0
            # Compute avg accuracy for this carrier
            accs = []
            for _, r in group.iterrows():
                rd = {}
                for c in df.columns:
                    v = r[c]
                    rd[c] = None if pd.isna(v) else v
                accs.append(_compute_row_accuracy(rd))
            avg_acc = sum(accs) / len(accs) if accs else 0
            methods.append({
                "carrier": str(carrier),
                "rows": carrier_rows,
                "mrc": round(carrier_mrc, 2),
                "avg_confidence": round(avg_acc, 1),
            })
        methods.sort(key=lambda x: x["rows"], reverse=True)

    return {
        "total_rows": total,
        "high": high,
        "medium": medium,
        "needs_review": needs_review,
        "high_pct": round(high / total * 100, 1) if total > 0 else 0,
        "medium_pct": round(medium / total * 100, 1) if total > 0 else 0,
        "needs_review_pct": round(needs_review / total * 100, 1) if total > 0 else 0,
        "high_mrc": round(high_mrc, 2),
        "medium_mrc": round(medium_mrc, 2),
        "needs_review_mrc": round(needs_review_mrc, 2),
        "extraction_methods": methods[:15],
    }


@router.get("/projects/{project_id}/inventory/filters")
async def inventory_filters(project_id: str, request: Request, source: str = "reference"):
    """Get unique filter values for dropdowns."""
    proj = request.app.state.projects.get(project_id)
    if not proj:
        return {"error": "Project not found"}

    file_path = _resolve_file(proj, source)
    df = load_inventory(file_path)
    if df.empty:
        return {}

    def _unique_sorted(col_name):
        col = next((c for c in df.columns if c.strip().lower() == col_name.lower()), None)
        if not col:
            col = next((c for c in df.columns if col_name.lower() in c.lower()), None)
        if not col:
            return []
        vals = df[col].dropna().astype(str).str.strip().unique().tolist()
        return sorted([v for v in vals if v and v != "nan"])

    return {
        "carriers": _unique_sorted("carrier"),
        "service_types": _unique_sorted("service type"),
        "charge_types": _unique_sorted("charge type"),
        "scu_codes": _unique_sorted("service or component"),
        "statuses": _unique_sorted("status"),
    }


@router.get("/projects/{project_id}/inventory/checklist")
async def get_checklist(project_id: str, request: Request, source: str = "reference"):
    """Get the checklist sheet data."""
    proj = request.app.state.projects.get(project_id)
    if not proj:
        return {"error": "Project not found"}

    file_path = _resolve_file(proj, source)
    if not file_path or not file_path.strip():
        return {"items": [], "columns": []}
    path = Path(file_path)
    if not path.exists() or not path.is_file():
        return {"items": [], "columns": []}

    # Find checklist sheet (may have leading space)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    checklist_sheet = None
    for name in wb.sheetnames:
        if "checklist" in name.lower():
            checklist_sheet = name
            break
    wb.close()

    if not checklist_sheet:
        return {"items": []}

    df = pd.read_excel(file_path, sheet_name=checklist_sheet, header=0)
    df.columns = [str(c).strip() for c in df.columns]

    items = []
    for _, row in df.iterrows():
        item = {}
        for col in df.columns:
            val = row[col]
            item[col] = None if pd.isna(val) else str(val)
        if item.get("Checklist"):  # Skip empty rows
            items.append(item)

    return {"items": items, "columns": list(df.columns)}


@router.post("/projects/{project_id}/inventory/checklist")
async def update_checklist(project_id: str, request: Request):
    """Update checklist items — persisted to disk."""
    from services.persistence import save_checklist
    data = await request.json()
    items = data.get("items", [])
    save_checklist(project_id, items)
    return {"status": "saved", "count": len(items)}


@router.post("/projects/{project_id}/inventory/row-status")
async def update_row_status(project_id: str, request: Request):
    """Update status and comment for a specific inventory row."""
    proj = request.app.state.projects.get(project_id)
    if not proj:
        return {"error": "Project not found"}

    data = await request.json()
    row_index = data.get("row_index")
    status = data.get("status", "in_progress")
    comment = data.get("comment", "")

    if row_index is None:
        return {"error": "row_index is required"}

    valid_statuses = ("completed", "need_review", "critical", "in_progress")
    if status not in valid_statuses:
        return {"error": f"Invalid status. Must be one of: {valid_statuses}"}

    if project_id not in _row_status_store:
        _row_status_store[project_id] = {}

    _row_status_store[project_id][int(row_index)] = {
        "status": status,
        "comment": comment,
    }

    return {"status": "saved", "row_index": row_index, "new_status": status}


@router.get("/projects/{project_id}/inventory/row-detail")
async def get_row_detail(
    project_id: str,
    request: Request,
    row_index: int = 0,
    source: str = "reference",
):
    """Get full detail for a single inventory row with all columns as field/value pairs."""
    proj = request.app.state.projects.get(project_id)
    if not proj:
        return {"error": "Project not found"}

    file_path = _resolve_file(proj, source)
    df = load_inventory(file_path)
    if df.empty or row_index not in df.index:
        return {"error": "Row not found"}

    row = df.loc[row_index]
    fields: list[dict] = []
    row_data: dict = {}

    for col in df.columns:
        val = row[col]
        if pd.isna(val):
            clean_val = None
        elif hasattr(val, "item"):
            clean_val = val.item()
        elif hasattr(val, "isoformat"):
            clean_val = val.isoformat()
        else:
            clean_val = val
        fields.append({"field_name": col, "field_value": clean_val})
        row_data[col] = clean_val

    accuracy = _compute_row_accuracy(row_data)
    status = _get_row_status(project_id, row_index, accuracy)
    store = _row_status_store.get(project_id, {})
    comment = store.get(row_index, {}).get("comment", "")
    # Get source files specific to THIS row
    source_files = _get_source_files(proj, row_data=row_data)

    return {
        "row_index": row_index,
        "fields": fields,
        "accuracy_score": accuracy,
        "status": status,
        "comment": comment,
        "source_documents": source_files,
    }


@router.get("/projects/{project_id}/inventory/export")
async def export_inventory(project_id: str, request: Request, source: str = "reference"):
    proj = request.app.state.projects.get(project_id)
    if not proj:
        return {"error": "Project not found"}

    if source == "extracted":
        output_dir = Path(proj["output_dir"])
        output_files = sorted(output_dir.glob("*_inventory_output.xlsx"))
        if output_files:
            return FileResponse(output_files[-1], filename=output_files[-1].name)

    ref_path = Path(proj["reference_file"])
    if ref_path.exists():
        return FileResponse(ref_path, filename=ref_path.name)

    return {"error": "No inventory file found"}


# ─── Auto-populate checklist ────────────────────────────

import re


def _find_col(df: pd.DataFrame, *names: str):
    """Find a DataFrame column matching any of the given names (case-insensitive partial match)."""
    for name in names:
        for col in df.columns:
            if name.lower() in col.lower():
                return col
    return None


def _check_s_record_missing(df: pd.DataFrame) -> dict:
    """Check whether S record missing and only C record exists."""
    scu_col = _find_col(df, "service or component")
    if not scu_col:
        return {"result": "N/A", "detail": "Column not found"}
    c_rows = df[df[scu_col].astype(str).str.strip() == "C"]
    s_rows = df[df[scu_col].astype(str).str.strip() == "S"]
    # Check if any C-row account IDs have no parent S-row
    acct_col = _find_col(df, "account", "acct")
    if not acct_col:
        return {"result": "N/A", "detail": "Account column not found"}
    s_accounts = set(s_rows[acct_col].dropna().astype(str).str.strip())
    orphan_c = c_rows[~c_rows[acct_col].astype(str).str.strip().isin(s_accounts)]
    count = len(orphan_c)
    return {
        "result": "Yes" if count > 0 else "No",
        "detail": f"{count} C-rows without parent S-row" if count > 0 else "All C-rows have parent S-rows",
    }


def _check_subtotal_mismatch(df: pd.DataFrame) -> dict:
    """Check if S-row MRC = sum of child C-row MRCs."""
    scu_col = _find_col(df, "service or component")
    mrc_col = _find_col(df, "monthly recurring")
    acct_col = _find_col(df, "account", "acct")
    if not all([scu_col, mrc_col, acct_col]):
        return {"result": "N/A", "detail": "Required columns not found"}

    s_rows = df[df[scu_col].astype(str).str.strip() == "S"].copy()
    c_rows = df[df[scu_col].astype(str).str.strip() == "C"].copy()
    s_rows[mrc_col] = pd.to_numeric(s_rows[mrc_col], errors="coerce").fillna(0)
    c_rows[mrc_col] = pd.to_numeric(c_rows[mrc_col], errors="coerce").fillna(0)

    mismatches = 0
    for _, s_row in s_rows.iterrows():
        acct = str(s_row[acct_col]).strip()
        s_mrc = round(float(s_row[mrc_col]), 2)
        children = c_rows[c_rows[acct_col].astype(str).str.strip() == acct]
        c_sum = round(float(children[mrc_col].sum()), 2)
        if abs(s_mrc - c_sum) > 0.01:
            mismatches += 1

    return {
        "result": "Yes" if mismatches > 0 else "No",
        "detail": f"{mismatches} accounts with sub-total mismatch" if mismatches > 0 else "All sub-totals match",
    }


def _check_blank_field(df: pd.DataFrame, *col_names: str) -> dict:
    """Check if any rows have blank values in the specified column."""
    col = _find_col(df, *col_names)
    if not col:
        return {"result": "N/A", "detail": f"Column not found: {col_names}"}
    blanks = df[col].isna() | (df[col].astype(str).str.strip() == "") | (df[col].astype(str).str.strip().str.lower() == "nan")
    count = int(blanks.sum())
    return {
        "result": "Yes" if count > 0 else "No",
        "detail": f"{count} rows with blank {col}" if count > 0 else f"All rows have {col} populated",
    }


def _check_phone_format(df: pd.DataFrame) -> dict:
    """Validate phone number formats."""
    phone_col = _find_col(df, "phone", "telephone", "btn", "billing telephone")
    if not phone_col:
        return {"result": "N/A", "detail": "Phone column not found"}
    phone_pattern = re.compile(r"^[\d\-\(\)\s\+\.]{7,20}$")
    invalid = 0
    for val in df[phone_col].dropna().astype(str):
        val = val.strip()
        if val and val.lower() != "nan" and not phone_pattern.match(val):
            invalid += 1
    return {
        "result": "Yes" if invalid > 0 else "No",
        "detail": f"{invalid} rows with invalid phone format" if invalid > 0 else "All phone numbers valid",
    }


def _check_duplicate_rows(df: pd.DataFrame) -> dict:
    """Check for duplicate rows based on key columns."""
    scu_col = _find_col(df, "service or component")
    acct_col = _find_col(df, "account", "acct")
    circuit_col = _find_col(df, "circuit", "service id")
    key_cols = [c for c in [acct_col, scu_col, circuit_col] if c]
    if not key_cols:
        return {"result": "N/A", "detail": "Key columns not found"}
    dupes = df.duplicated(subset=key_cols, keep=False)
    count = int(dupes.sum())
    return {
        "result": "Yes" if count > 0 else "No",
        "detail": f"{count} potential duplicate rows" if count > 0 else "No duplicates found",
    }


# Map of checklist text keywords to validation functions
_CHECKLIST_VALIDATORS = {
    "s record missing": _check_s_record_missing,
    "only c record": _check_s_record_missing,
    "sub total mismatch": _check_subtotal_mismatch,
    "subtotal mismatch": _check_subtotal_mismatch,
    "service address not available": lambda df: _check_blank_field(df, "service address", "address"),
    "service address": lambda df: _check_blank_field(df, "service address", "address"),
    "billing name": lambda df: _check_blank_field(df, "billing name", "billing account name"),
    "billing names should not be blank": lambda df: _check_blank_field(df, "billing name", "billing account name"),
    "phone number format": _check_phone_format,
    "phone number": _check_phone_format,
    "duplicate": _check_duplicate_rows,
    "contract expiry": lambda df: _check_blank_field(df, "contract", "expiry", "term"),
    "status blank": lambda df: _check_blank_field(df, "status"),
    "carrier blank": lambda df: _check_blank_field(df, "carrier"),
    "mrc zero": lambda df: {
        "result": "Yes" if int((pd.to_numeric(df[_find_col(df, "monthly recurring")] if _find_col(df, "monthly recurring") else pd.Series(), errors="coerce").fillna(0) == 0).sum()) > 0 else "No",
        "detail": f"{int((pd.to_numeric(df[_find_col(df, 'monthly recurring')] if _find_col(df, 'monthly recurring') else pd.Series(), errors='coerce').fillna(0) == 0).sum())} rows with zero MRC",
    },
}


@router.post("/projects/{project_id}/inventory/checklist/auto-populate")
async def auto_populate_checklist(project_id: str, request: Request):
    """Auto-populate checklist based on extracted data validation."""
    proj = request.app.state.projects.get(project_id)
    if not proj:
        return {"error": "Project not found"}

    # Load extracted data (prefer extracted, fall back to reference)
    file_path = _resolve_file(proj, "extracted")
    df = load_inventory(file_path)
    if df.empty:
        file_path = _resolve_file(proj, "reference")
        df = load_inventory(file_path)

    if df.empty:
        return {"error": "No inventory data available"}

    # Load current checklist
    ref_path = proj.get("reference_file", "")
    checklist_items = []
    if ref_path and Path(ref_path).exists():
        wb = openpyxl.load_workbook(ref_path, read_only=True, data_only=True)
        checklist_sheet = None
        for name in wb.sheetnames:
            if "checklist" in name.lower():
                checklist_sheet = name
                break
        wb.close()

        if checklist_sheet:
            cdf = pd.read_excel(ref_path, sheet_name=checklist_sheet, header=0)
            cdf.columns = [str(c).strip() for c in cdf.columns]
            for _, row in cdf.iterrows():
                item = {}
                for col in cdf.columns:
                    val = row[col]
                    item[col] = None if pd.isna(val) else str(val)
                if item.get("Checklist"):
                    checklist_items.append(item)

    if not checklist_items:
        return {"error": "No checklist items found"}

    # Auto-populate each checklist item
    results = []
    for item in checklist_items:
        text = (item.get("Checklist") or "").lower()
        validation_result = None

        # Try to match against known validators
        for keyword, validator in _CHECKLIST_VALIDATORS.items():
            if keyword in text:
                try:
                    validation_result = validator(df)
                except Exception:
                    validation_result = {"result": "N/A", "detail": "Validation error"}
                break

        if validation_result:
            item["Agent - Yes/No"] = validation_result["result"]
            item["_auto_detail"] = validation_result["detail"]
        else:
            item["Agent - Yes/No"] = item.get("Agent - Yes/No") or "N/A"
            item["_auto_detail"] = "No automated check available"

        results.append(item)

    return {
        "items": results,
        "auto_populated": sum(1 for r in results if r.get("_auto_detail") != "No automated check available"),
        "total": len(results),
    }
